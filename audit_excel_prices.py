from openpyxl import load_workbook

EXCEL_PATH = r'C:\app precificador\precificador-sfimports\SF_IMPORTS_Pontuacoes_v4_FINAL.xlsx'
wb = load_workbook(EXCEL_PATH, data_only=True)
sheet = wb['🍷 Todos os Produtos']

total_names = 0
with_milao_price = 0
with_sf_price = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    name = row[1]
    if name:
        total_names += 1
        milao = row[3]
        sf = row[5]
        
        try:
            if milao and float(milao) > 0:
                with_milao_price += 1
        except: pass
        
        try:
            if sf and float(sf) > 0:
                with_sf_price += 1
        except: pass

print(f"Excel Audit:")
print(f" - Names found: {total_names}")
print(f" - Rows with Milao Por > 0: {with_milao_price}")
print(f" - Rows with SF Final > 0: {with_sf_price}")
