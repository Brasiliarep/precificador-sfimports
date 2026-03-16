from openpyxl import load_workbook
import re

def parse_price(val):
    if val is None or val == "" or val == "—": return 0
    if isinstance(val, (int, float)): return float(val)
    # Remove R$, dots and replace comma with dot
    s = str(val).replace('R$', '').replace('.', '').replace(',', '.').strip()
    try:
        return float(s)
    except:
        return 0

EXCEL_PATH = r'C:\app precificador\precificador-sfimports\SF_IMPORTS_Pontuacoes_v4_FINAL.xlsx'
wb = load_workbook(EXCEL_PATH, data_only=True)
sheet = wb['🍷 Todos os Produtos']

names_total = 0
milao_prices = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    name = row[1]
    if name:
        names_total += 1
        price = parse_price(row[3])
        if price > 0:
            milao_prices += 1

print(f"Final Excel Audit:")
print(f" - Total Names: {names_total}")
print(f" - Valid Milao Prices Found: {milao_prices}")
